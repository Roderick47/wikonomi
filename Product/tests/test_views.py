from os import remove
from django.test import TestCase, Client
from django.urls import reverse
from Photo.forms import ProductPhotoAddForm
from Product.forms import GetOrCreateBusinessForm, ProductAddForm
from Product.models import Product
from Business.models import Business
from Location.models import Location
from django.contrib.auth.models import User
from urllib.parse import urlencode
import tempfile
import csv
from openpyxl import Workbook

class TestViews(TestCase):

    def setUp(self ):
        self.client = Client()
        # Create test user first
        self.test_user = User.objects.create_user(username="test_user", password="test_password")
        
        # Create test location
        self.test_location = Location.objects.create(
            latitude=0.0,
            longitude=0.0
        )
        # Create test business with Location instance and author
        self.test_business = Business.objects.create(
            name = 'test_business',
            description = 'business_description',
            location = self.test_location,
            author = self.test_user,  # Add author
        )
        self.test_product = Product.objects.create(
            name="test_product",
            price=10,
            description="test_description",
            business=self.test_business,
            author=self.test_user,
            )
        self.file=tempfile.NamedTemporaryFile()

    # Tests for ProductAddView
    def test_ProductAddView__GET__not_logged_in(self):
        url = reverse('Product:add',kwargs={"bus_id":1})
        response = self.client.get(url)
        redirect_url = '{}?{}'.format(reverse("account_login"), urlencode({"next":url}))
        self.assertRedirects(response,redirect_url,302,200)

    def test_ProductAddView__GET__logged_in(self):
        url = reverse('Product:add',kwargs={"bus_id":1})
        self.client.force_login(self.test_user)
        response = self.client.get(url)
        self.assertEquals(response.status_code,200)
        self.assertTemplateUsed(response,'Product/ProductAddForm.html')
        self.assertTrue(isinstance(response.context['form'],ProductAddForm))
        self.assertTrue(isinstance(response.context['imageForm'],ProductPhotoAddForm))
        self.assertTrue(isinstance(response.context['business'],Business))

    def test_ProductAddView__POST__no_data(self):
        url = reverse('Product:add',kwargs={"bus_id":1})
        self.client.force_login(self.test_user)
        response=self.client.post(url,{})
        self.assertEquals(response.status_code,200)
        self.assertTemplateUsed(response,'Product/ProductAddForm.html')
        self.assertTrue(isinstance(response.context['form'],ProductAddForm))
        self.assertTrue(isinstance(response.context['imageForm'],ProductPhotoAddForm))
        self.assertTrue(isinstance(response.context['business'],Business))

    def test_ProductAddView__POST__with_data(self):
        url = reverse('Product:add',kwargs={"bus_id":1})
        self.client.force_login(self.test_user)
        response=self.client.post(url,{
            "name":"test_product_edited",
            "price":50,
            "description":"description_edited",
            "photo":self.file.name,
            "latitude":1.0,
            "longitude":1.0,
            "use_browser_location":True
            })
        redirect_url = reverse("Product:detail",kwargs={"prod_id":Product.objects.last().id})
        self.assertRedirects(response,redirect_url,302,200)
        self.assertEqual(len(Product.objects.all()),2)


    # Tests for ProductDetailView:
    # Authentication not needed for this view to render product details to user.
    def test_ProductDetailView__GET__not_Logged_in(self):
        url = reverse("Product:detail",kwargs={"prod_id":1})
        response = self.client.get(url)
        self.assertEquals(response.status_code,200)
        self.assertTemplateUsed(response,'Product/ProductDetail.html')

    # Tests for ProductEditView:

    def test_ProductEditView__GET__not_logged_in(self):
        url = reverse("Product:edit",kwargs={"prod_id":1})
        response = self.client.get(url)
        redirect_url = '{}?{}'.format(reverse("account_login"), urlencode({"next":url}))
        self.assertEquals(response.status_code,302)
        self.assertRedirects(response,redirect_url,302,200)

    def test_ProductEditView__GET__logged_in__wrong_Product_id(self):
        self.client.force_login(self.test_user)
        url = reverse("Product:edit",kwargs={"prod_id":100})
        response = self.client.get(url)
        self.assertEquals(response.status_code,404)

    def test_ProductEditView__GET__logged_in__right_Product_id(self):
        self.client.force_login(self.test_user)
        url = reverse("Product:edit",kwargs={"prod_id":self.test_product.id})
        response = self.client.get(url)
        self.assertEquals(response.status_code,200)
        self.assertTemplateUsed(response,'Product/ProductEditForm.html')
        self.assertTrue(isinstance(response.context['product'],Product))
        self.assertTrue(isinstance(response.context['form'],ProductAddForm))
        self.assertTrue(isinstance(response.context['imageForm'],ProductPhotoAddForm))

    def test_ProductEditView__POST__logged_in__right_product_id__no_data(self):
        self.client.force_login(self.test_user)
        url = reverse("Product:edit",kwargs={"prod_id":self.test_product.id})
        response = self.client.post(url,kwargs={})
        self.assertEquals(response.status_code,200)
        self.assertEquals(len(Product.objects.all()),1)
        self.assertTemplateUsed(response,'Product/ProductEditForm.html')
        self.assertTrue(isinstance(response.context['product'],Product))
        self.assertTrue(isinstance(response.context['form'],ProductAddForm))
        self.assertTrue(isinstance(response.context['imageForm'],ProductPhotoAddForm))

    def test_ProductEditView__POST__logged_in__right_product_id__with_data(self):
        self.client.force_login(self.test_user)
        url = reverse("Product:edit",kwargs={"prod_id":self.test_product.id})
        response = self.client.post(url,{
            "name":"test_product_edited",
            "price":50,
            "description":"description_edited",
            "photo":self.file.name,
            "latitude":1.0,
            "longitude":1.0,
            "use_browser_location":True
            })
        self.assertEqual(len(Product.objects.all()),1)
        self.assertEqual(Product.objects.first().name,"test_product_edited")
        self.assertEquals(response.status_code,302)
        redirect_url = reverse("Product:detail",args=[Product.objects.first().id])
        self.assertRedirects(response,redirect_url,302,200)

    # Test for BusinessProductListView
    def test_BusinessProductListView__GET__not_logged_in(self):
        url = reverse("Product:business-list",kwargs={"bus_id":1})
        response = self.client.get(url)
        self.assertEquals(response.status_code,200)
        self.assertTemplateUsed(response,'Product/BusinessProductGallery.html')
        self.assertTrue(isinstance(response.context['business'],Business))

    # Tests for  ProductAddGenerlView:
    def test_ProductAddGeneralView__GET__not_logged_in(self):
        url = reverse("Product:add-general")
        response = self.client.get(url)
        redirect_url = '{}?{}'.format(reverse("account_login"), urlencode({"next":url}))
        self.assertEquals(response.status_code, 302)
        self.assertRedirects(response,redirect_url,302,200)

    def test_ProductAddGeneralView__POST__not_logged_in(self):
        url=reverse("Product:add-general")
        response = self.client.post(url)
        redirect_url = '{}?{}'.format(reverse("account_login"), urlencode({"next":url}))
        self.assertEquals(response.status_code,302)
        self.assertRedirects(response,redirect_url,302,200)

    def test_ProductAddGeneralView__GET__logged_in(self):
        url = reverse("Product:add-general")
        self.client.force_login(self.test_user)
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        self.assertTemplateUsed(response,'Product/GeneralProductAddForm.html')
        self.assertTrue(isinstance(response.context['form'],ProductAddForm))
        self.assertTrue(isinstance(response.context['form2'],GetOrCreateBusinessForm))
        self.assertTrue(isinstance(response.context['imageForm'],ProductPhotoAddForm))

    def test_ProductAddGeneralView__POST__logged_in__with_data(self):
        new_business = Business.objects.create(
            name="test_business2",
            description="test_description",
            location=self.test_location
        )
        url = reverse("Product:add-general")
        self.client.force_login(self.test_user)
        response = self.client.post(url,{
            "name": "test_product_name",
            "price": 12,
            "business": "test_business2",
            "image":self.file.name,
            "latitude":1.0,
            "longitude":1.0,
            "use_browser_location":True
            })
        self.assertEquals(response.status_code, 302)
        self.assertEqual(len(Product.objects.all()),2)

    def test_ProductAddGeneralView__POST__logged_in__no_data(self):
        url = reverse("Product:add-general")
        self.client.force_login(self.test_user)
        response = self.client.post(url,kwargs={})
        self.assertEquals(response.status_code, 200)
        self.assertEqual(len(Product.objects.all()),1)
        self.assertEqual(len(Business.objects.all()),1)
        self.assertTemplateUsed(response,'Product/GeneralProductAddForm.html')
    
    # Tests for ProductDeleteView:
    def test_ProductDeleteView__GET__not_logged_in(self):
        url = reverse("Product:delete",kwargs={"prod_id":self.test_product.id})
        response = self.client.get(url)
        redirect_url= reverse("Product:detail",kwargs={"prod_id":self.test_product.id})
        self.assertRedirects(response,redirect_url,302,200)

    def test_ProductDeleteView__GET__logged_in(self):
        url = reverse("Product:delete",kwargs={"prod_id":self.test_product.id})
        self.client.force_login(self.test_user)
        response = self.client.get(url)
        self.assertEqual(len(Product.objects.all()),0)
        redirect_url= reverse("Home:home")
        self.assertRedirects(response,redirect_url,302,200)

    # Tests for ProductListTemplateDownload
    def test_ProductListTemplateDownload__GET__not_logged_in(self):
        url=reverse("Product:template-download")
        response = self.client.get(url)
        self.assertEquals(response.get('Content-Disposition'),"attachment; filename=file_download/product_list_template.xlsx")

    # Test ProductListUploadView
    def test_ProductListUploadView__GET__not_logged_in(self):
        url=reverse("Product:list-upload",kwargs={"bus_id":1})
        response = self.client.get(url)
        redirect_url = '{}?{}'.format(reverse("account_login"), urlencode({"next":url}))
        self.assertRedirects(response,redirect_url,302,200)
        
    def test_ProductListUploadView__GET__logged_in(self):
        url=reverse("Product:list-upload",kwargs={"bus_id":1})
        self.client.force_login(self.test_user)
        response = self.client.get(url)
        self.assertEquals(response.status_code, 200)
        self.assertTemplateUsed(response,'Product/import_product_list.html')

    def test_ProductListUploadView__POST__no_data(self):
        url=reverse("Product:list-upload",kwargs={"bus_id":1})
        self.client.force_login(self.test_user)
        response = self.client.post(url,kwargs={"myfile":""})
        self.assertEquals(response.status_code, 200)
        self.assertTemplateUsed(response,'Product/import_product_list.html')
        self.assertEqual(len(Product.objects.all()),1)

    def test_ProductListUploadView__POST__with_data(self):
        """Test successful product upload with valid Excel file"""
        url=reverse("Product:list-upload",kwargs={"bus_id":self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create a proper Excel file using openpyxl
        file_name = "test.xlsx"
        wb = Workbook()
        ws = wb.active
        # Add headers
        ws.append(['name', 'price', 'description'])
        # Add data rows
        ws.append(['Test_Product2', 12, 'test_description'])
        ws.append(['Test_Product3', 30, '3test_description'])
        wb.save(file_name)
        
        # Upload the Excel file
        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})
        
        redirect_url = reverse("Product:business-list", kwargs={"bus_id": self.test_business.id})
        self.assertRedirects(response, redirect_url, 302, 200)
        self.assertEqual(len(Product.objects.all()), 3)
        remove(file_name)
    
    def test_ProductListUploadView__POST__unauthorized_user(self):
        """Test that users cannot upload to private businesses they don't own"""
        # Create a private business owned by test_user
        private_business = Business.objects.create(
            name='private_business',
            description='private business',
            location=self.test_location,
            author=self.test_user,
            is_public=False  # Make it private
        )

        # Create another user
        other_user = User.objects.create_user(username="other_user", password="testpass")

        url = reverse("Product:list-upload", kwargs={"bus_id": private_business.id})
        self.client.force_login(other_user)

        # Create Excel file
        file_name = "test_unauthorized.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['name', 'price', 'description'])
        ws.append(['Unauthorized_Product', 100, 'Should not be created'])
        wb.save(file_name)

        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})

        # Should redirect to business detail with error, not create products
        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(Product.objects.all()), 1)  # Only the original test product
        remove(file_name)
    
    def test_ProductListUploadView__POST__authorized_user_public_business(self):
        """Test that users can upload to public businesses they don't own"""
        # Create another user
        other_user = User.objects.create_user(username="other_user", password="testpass")

        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(other_user)

        # Create Excel file
        file_name = "test_authorized.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['name', 'price', 'description'])
        ws.append(['Authorized_Product', 150, 'Should be created by other user'])
        wb.save(file_name)

        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})

        # Should succeed since test_business is public
        redirect_url = reverse("Product:business-list", kwargs={"bus_id": self.test_business.id})
        self.assertRedirects(response, redirect_url, 302, 200)
        self.assertEqual(len(Product.objects.all()), 2)  # Original + new product
        remove(file_name)
    
    def test_ProductListUploadView__POST__invalid_file_type(self):
        """Test that non-Excel files are rejected"""
        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create a text file instead of Excel
        file_name = "test.txt"
        with open(file_name, "w") as f:
            f.write("This is not an Excel file")
        
        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})
        
        # Should stay on same page with error
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'Product/import_product_list.html')
        self.assertEqual(len(Product.objects.all()), 1)  # No new products created
        remove(file_name)
    
    def test_ProductListUploadView__POST__file_too_large(self):
        """Test that files exceeding size limit are rejected"""
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create a mock file that's too large (> 5MB)
        large_file = SimpleUploadedFile(
            "large_file.xlsx",
            b"x" * (6 * 1024 * 1024),  # 6MB
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(url, {"myfile": large_file})
        
        # Should stay on same page with error
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(Product.objects.all()), 1)  # No new products created
    
    def test_ProductListUploadView__POST__negative_price(self):
        """Test that products with negative prices are rejected"""
        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create Excel with negative price
        file_name = "test_negative.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['name', 'price', 'description'])
        ws.append(['Bad_Product', -10, 'Negative price'])
        wb.save(file_name)
        
        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})
        
        # Should stay on same page with validation error
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(Product.objects.all()), 1)  # No new products created
        remove(file_name)
    
    def test_ProductListUploadView__POST__empty_file(self):
        """Test that empty Excel files are rejected"""
        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create empty Excel file
        file_name = "test_empty.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['name', 'price', 'description'])  # Headers only, no data
        wb.save(file_name)
        
        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})
        
        # Should stay on same page with error
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(Product.objects.all()), 1)  # No new products created
        remove(file_name)
    
    def test_ProductListUploadView__POST__xss_attempt(self):
        """Test that XSS attempts in product names are sanitized"""
        url = reverse("Product:list-upload", kwargs={"bus_id": self.test_business.id})
        self.client.force_login(self.test_user)
        
        # Create Excel with XSS attempt
        file_name = "test_xss.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['name', 'price', 'description'])
        ws.append(['<script>alert("XSS")</script>', 100, '<img src=x onerror=alert(1)>'])
        wb.save(file_name)
        
        with open(file_name, "rb") as file_data:
            response = self.client.post(url, {"myfile": file_data})
        
        # Should succeed but sanitize the data
        self.assertEqual(response.status_code, 302)
        new_product = Product.objects.last()
        # Check that script tags are removed
        self.assertNotIn('<script>', new_product.name)
        self.assertNotIn('<img', new_product.description)
        remove(file_name)

    # def test_ProductAutoComplete(self):
    #     url = reverse("Product:autocomplete")
    #     response = self.client.get(url)

